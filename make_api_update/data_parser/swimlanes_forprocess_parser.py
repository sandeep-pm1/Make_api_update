from openpyxl import load_workbook
from fastapi import HTTPException
import os

#parsing the xlsx file and creating dic with stage as key and swimlanes as list of values
def create_dict_from_excel(file_path):
    #checking for file existence if not raise the exception
    if(not os.path.isfile(file_path)):
        raise HTTPException(404, detail="Data not Available")
    else:
        wb = load_workbook(file_path)
        sheet = wb.active

        data_dict = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = row[0]
            value = row[1]

            if key not in data_dict:
                data_dict[key] = []

            data_dict[key].append(value)

    return data_dict

